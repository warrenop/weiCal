"""Parse WeChat Pay / Alipay monthly bill (CSV or XLSX) and insert into the
database with dedup.

The CSV layout depends on the source:

- **WeChat**: 16-line preamble, header row starts with `交易时间`.
- **Alipay**: ~25-line preamble, header row starts with `交易号`. Columns are
  named differently; we map them onto our normalized row schema.

`parse_bill()` auto-detects which source produced the file and dispatches.
"""
import csv
import io
import re
from typing import Callable, Iterable, Optional

from openpyxl import load_workbook

from .categorizer import categorize
from .db import get_conn


# ---------- source profiles ----------
#
# Each profile knows:
#   - `header_marker`: the literal cell that starts the real header row, used
#     to skip the preamble.
#   - `extract`: maps a parsed dict (cleaned header → value) to our normalized
#     row, returning None if the row should be skipped.

def _to_iso(s: str) -> str:
    return (s or "").strip().replace("/", "-")


def _amount(raw) -> float:
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = re.sub(r"[¥￥,\s]", "", str(raw))
    return float(s) if s else 0.0


def _wechat_direction(s: str) -> str:
    s = (s or "").strip()
    return {"支出": "expense", "收入": "income"}.get(s, "neutral")


def _alipay_direction(s: str) -> str:
    s = (s or "").strip()
    if s == "支出":
        return "expense"
    if s == "收入":
        return "income"
    # "不计收支" → inter-account transfers; not interesting for spending analysis.
    # Return special sentinel; extractor will skip.
    if s == "不计收支":
        return ""
    return "neutral"


def _extract_wechat(r: dict) -> Optional[dict]:
    tx_time = _to_iso(str(r.get("交易时间", "") or ""))
    if not tx_time:
        return None
    direction = _wechat_direction(r.get("收/支", ""))
    counterparty = r.get("交易对方", "") or ""
    product = r.get("商品", "") or ""
    tx_type = r.get("交易类型", "") or ""
    return {
        "tx_time": tx_time,
        "tx_type": tx_type,
        "counterparty": counterparty,
        "product": product,
        "amount": _amount(r.get("金额(元)") or r.get("金额") or "0"),
        "direction": direction,
        "pay_method": r.get("支付方式", "") or "",
        "status": r.get("当前状态", "") or "",
        "wx_tx_id": (r.get("交易单号", "") or "") or None,
        "category": categorize(counterparty, product, tx_type, direction),
        "source": "wechat_csv",
        "notes": r.get("备注", "") or "",
        "period": tx_time[:7],
    }


def _extract_alipay(r: dict) -> Optional[dict]:
    tx_time = _to_iso(str(r.get("交易创建时间", "") or r.get("付款时间", "") or ""))
    if not tx_time:
        return None
    direction = _alipay_direction(r.get("收/支", ""))
    if not direction:
        return None  # skip 不计收支
    counterparty = (r.get("交易对方", "") or "").strip()
    product = (r.get("商品名称", "") or r.get("商品", "") or "").strip()
    tx_type = (r.get("类型", "") or "").strip()
    status = (r.get("交易状态", "") or "").strip()
    if status in ("等待付款", "交易关闭", "已关闭"):
        return None  # failed / cancelled
    # Alipay's 交易来源地 (淘宝 / 天猫 / 飞猪 / ...) is a strong category signal —
    # fold it into the text the categorizer sees.
    origin = (r.get("交易来源地", "") or "").strip()
    return {
        "tx_time": tx_time,
        "tx_type": tx_type,
        "counterparty": counterparty,
        "product": product,
        "amount": _amount(r.get("金额（元）") or r.get("金额(元)") or r.get("金额") or "0"),
        "direction": direction,
        "pay_method": (r.get("资金状态", "") or "支付宝").strip(),
        "status": status,
        "wx_tx_id": ((r.get("交易号", "") or r.get("流水号", "") or "").strip()) or None,
        "category": categorize(counterparty, f"{product} {origin}", tx_type, direction),
        "source": "alipay_csv",
        "notes": (r.get("备注", "") or "").strip(),
        "period": tx_time[:7],
    }


PROFILES = {
    "wechat": {"header_marker": "交易时间", "extract": _extract_wechat},
    "alipay": {"header_marker": "交易号", "extract": _extract_alipay},
}


# ---------- low-level CSV / XLSX scanners ----------

def _decode(raw_bytes: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return raw_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法解码文件（尝试了 utf-8 / gbk / gb18030）")


def _detect_source(lines: list[str]) -> tuple[Optional[str], int]:
    """Return (source_key, header_offset). source_key is None if unknown."""
    for i, ln in enumerate(lines):
        # Alipay header lines may have leading commas / spaces around tokens.
        first_cell = ln.split(",", 1)[0].strip()
        if first_cell.startswith(PROFILES["wechat"]["header_marker"]):
            return ("wechat", i)
        if first_cell.startswith(PROFILES["alipay"]["header_marker"]):
            return ("alipay", i)
    return (None, 0)


def _parse_csv(raw_bytes: bytes) -> tuple[str, Iterable[dict]]:
    text = _decode(raw_bytes)
    lines = text.splitlines()
    source, offset = _detect_source(lines)
    if not source:
        raise ValueError(
            "无法识别账单类型（既不像微信也不像支付宝的标准 CSV）。"
            "请确认下载的是官方账单原始文件。"
        )
    body = "\n".join(lines[offset:])
    reader = csv.DictReader(io.StringIO(body))
    return source, reader


def _parse_xlsx(raw_bytes: bytes) -> tuple[str, Iterable[dict]]:
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    header: list[str] | None = None
    source: Optional[str] = None
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else (c if isinstance(c, str) else str(c)) for c in row]
        if header is None:
            first = cells[0].strip() if cells else ""
            if first.startswith(PROFILES["wechat"]["header_marker"]):
                source = "wechat"
                header = [c.strip() for c in cells]
            elif first.startswith(PROFILES["alipay"]["header_marker"]):
                source = "alipay"
                header = [c.strip() for c in cells]
            continue
        if not any(cells):
            continue
        rows.append(dict(zip(header, cells)))
    if not source:
        raise ValueError("xlsx 中未找到熟悉的账单表头")
    return source, iter(rows)


def _looks_like_xlsx(raw_bytes: bytes, file_name: str) -> bool:
    if file_name.lower().endswith((".xlsx", ".xlsm")):
        return True
    return raw_bytes[:4] == b"PK\x03\x04"


def parse_bill(raw_bytes: bytes, file_name: str = "") -> tuple[str, list[dict]]:
    """Return (source, normalized_rows)."""
    source, raw_rows = (
        _parse_xlsx(raw_bytes) if _looks_like_xlsx(raw_bytes, file_name)
        else _parse_csv(raw_bytes)
    )
    extract = PROFILES[source]["extract"]
    # Header keys may have trailing/leading whitespace; normalize.
    rows: list[dict] = []
    for r in raw_rows:
        cleaned = {
            (k or "").strip(): (v.strip() if isinstance(v, str) else v)
            for k, v in r.items()
        }
        n = extract(cleaned)
        if n is not None:
            rows.append(n)
    return source, rows


def import_bill(raw_bytes: bytes, file_name: str) -> dict:
    source, rows = parse_bill(raw_bytes, file_name)
    inserted = skipped = failed = 0
    period_start = period_end = None
    if rows:
        times = sorted(r["tx_time"] for r in rows)
        period_start, period_end = times[0][:10], times[-1][:10]

    with get_conn() as conn:
        for row in rows:
            try:
                cur = conn.execute(
                    """INSERT OR IGNORE INTO transactions
                       (tx_time, tx_type, counterparty, product, amount, direction,
                        pay_method, status, wx_tx_id, category, source, notes, period)
                       VALUES (:tx_time, :tx_type, :counterparty, :product, :amount, :direction,
                               :pay_method, :status, :wx_tx_id, :category, :source, :notes, :period)""",
                    row,
                )
                if cur.rowcount == 1:
                    inserted += 1
                else:
                    skipped += 1
            except Exception:
                failed += 1
        conn.execute(
            """INSERT INTO import_logs
               (file_name, period_start, period_end, inserted, skipped, failed, source)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (file_name, period_start, period_end, inserted, skipped, failed, source),
        )

    return {
        "source": source,
        "inserted": inserted,
        "skipped": skipped,
        "failed": failed,
        "period_start": period_start,
        "period_end": period_end,
        "total": len(rows),
    }


# Back-compat aliases for older route names + tests
parse_wechat_csv = parse_wechat_bill = parse_bill
import_wechat_csv = import_wechat_bill = import_bill
